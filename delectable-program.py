#-*-encoding:utf8-*-
from bs4 import BeautifulSoup
from openpyxl import load_workbook,Workbook
import os
from selenium import webdriver

def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        sheet.append(_DATA)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = '시트이름'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 50
        book.save(_FILENAME)

def log(tag, text):
	# Info tag
	if(tag == 'i'):
		print("[INFO] " + text)
	# Error tag
	elif(tag == 'e'):
		print("[ERROR] " + text)
	# Success tag
	elif(tag == 's'):
		print("[SUCCESS] " + text)
'''----------------- 환경설정 부분 ---------------------'''
HEADER = ['Brand','Name', 'Rating', 'Region', 'Varietal', 'Pairings']
if __name__ == "__main__":
    FILENAME = input(">>> 저장 엑셀파일이름을 확장자포함(.xlsx) 적어주세요 :: ").strip()
    save_excel(FILENAME, None, HEADER)
    driver = webdriver.Chrome('./chromedriver')
    driver.maximize_window()
    driver.get('https://delectable.com/')
    while True:
        inputStr = input(">>> 추출 원하는 페이지에 들어가서 아무키나 눌러주세요 (종료는 0) :").strip()
        if inputStr == '0':
            break
        dataList = []
        bs4 = BeautifulSoup(driver.page_source, 'lxml')

        brand = bs4.find('h2', class_='wine-profile-header__producer').get_text().strip()
        name = bs4.find('h1', class_='wine-profile-header__name').get_text().strip()
        ratingValue = bs4.find("span", itemprop='ratingValue').get_text().strip()
        ratingCount = bs4.find("span", class_='wine-profile-rating__count').get_text().strip()
        proValue = bs4.find_all('div', class_='wine-profile-rating__rating')[1].find("span").get_text().strip()
        proCount = bs4.find_all('div', class_='wine-profile-rating__rating')[1].find("span",
                                                                                     class_='wine-profile-rating__count').get_text().strip()
        try:
            regionName = bs4.find('span', class_='wine-profile-region__name').get_text().strip()
        except:
            regionName = '-'
        try:
            varietalName = bs4.find('span', class_='wine-profile-varietal__name').get_text().strip()
        except:
            varietalName = '-'
        try:
            pairings = bs4.find('div', class_='wine-profile-pairings__name').get_text().strip()
        except:
            pairings = '-'
        dataList.append(brand)
        dataList.append(name)
        dataList.append(ratingValue + ' ' + ratingCount + ',' + proValue + ' ' + proCount)
        dataList.append(regionName)
        dataList.append(varietalName)
        dataList.append(pairings)
        save_excel(FILENAME, dataList, None)
        log('s',name+" 저장 완료")
    driver.quit()