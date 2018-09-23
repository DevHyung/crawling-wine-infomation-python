import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook,Workbook
import os
def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        #sheet.append(_DATA)
        for depth1List in _DATA:
            sheet.append(depth1List)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = '시트이름'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40
        book.save(_FILENAME)

header = {
'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36',
'x-requested-with': 'XMLHttpRequest',
'content-type': 'application/json'
}
'''----------------- 환경설정 부분 ---------------------'''
FILENAME = input(">>> 저장할 파일이름만 적어주세요(확장자미포함) :")+'.xlsx'
HEADER = ['Winery','Name', 'Region', 'Country', 'Rating', 'Summary']
# init
save_excel(FILENAME,None,HEADER)
f2 = open('예외/'+FILENAME+'_except.txt','a')
'''---------------------------------------------------'''
# 1 ~ 500  = 0번부터 499 까지
# 500 ~ 1000 = 499부터
#
print(">>> EX )  시작 1, 끝 5을  1번째부터 5번째까지 파싱합니다.")
START = int(input(">>> 시작 URL 번호(행) 입력 :")) - 1 # 715550
END = int(input(">>> 끝 URL 번호(행) 입력 : "))        # 737780
lines = open('vivino_url_list.txt').readlines()[START:END]
idx = START + 1
tmp = []
for url in lines:
    print("{}번째 : {} 실행중..".format(idx, url.strip()))
    idx += 1
    dataList = []
    #
    try:
        html = requests.get(url.strip(),timeout=20,headers=header)
    except:
        print('@@@')
        continue
    bs4 = BeautifulSoup(html.text,'lxml')
    #
    try:
        name = bs4.find('span',class_='wine-page__header__information__details__name__vintage').get_text().strip()
    except:#새로운페이지
        name = '-'
        print("\t>>> Move Except File")
        f2.write(url)
        continue
    try:
        winery = bs4.find('a',class_='wine-page__header__information__details__name__winery').get_text().strip()
    except:
        winery = '-'
    try:
        region = bs4.find('a',class_='wine-page__header__information__details__location__region').get_text().strip()
    except:
        region = '-'
    try:
        country = bs4.find("a",class_='wine-page__header__information__details__location__country').get_text().strip()
    except:
        country = '-'
    try:
        averLabel = bs4.find('div',class_='wine-page__header__information__details__average-rating__label').get_text().strip()
    except:
        averLabel = '-'
    try:
        averValue = bs4.find('div',class_='wine-page__header__information__details__average-rating__value__number').get_text().strip()
    except:
        averValue = '-'
    divs = bs4.find_all('div',class_='wine-page__summary__item')
    summary = ""
    for div in divs:
        summaryHeader = div.find('div',class_='wine-page__summary__item__header').get_text().strip()
        summary += summaryHeader+":"
        isVoid = True
        for a in div.find_all('a'):
            summary +=a.get_text().strip()+','
            isVoid = False
        for a in div.find_all('span'):
            summary += a.get_text().strip() + ','
            isVoid = False
        if isVoid:
            summary += div.find('div',class_='wine-page__summary__item__content').get_text().strip() + ','
        summary = summary[:-1]+'\n'

    dataList.append(winery)
    dataList.append(name)
    dataList.append(region)
    dataList.append(country)
    dataList.append(averLabel+' '+averValue)
    dataList.append(summary)
    tmp.append(dataList)
    if len(tmp) == 100:
        save_excel(FILENAME,tmp,None)
        tmp.clear()
save_excel(FILENAME,tmp,None)
f2.close()