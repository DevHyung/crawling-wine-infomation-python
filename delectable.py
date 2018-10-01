import requests
from bs4 import BeautifulSoup
from selenium import webdriver
import os
from openpyxl import load_workbook,Workbook
import time

def get_bs_obejct_by_url(url):
    html = requests.get(url)
    # print(html.encoding) # ISO-8859-1 인코딩나와서
    #html.encoding = 'euc-kr'  # 한글 인코딩으로 변환
    return BeautifulSoup(html.text, 'lxml')

def log(tag, text):
	# Info tag
	if(tag == 'i'):
		print("[INFO] " + text)
	# Error tag
	elif(tag == 'e'):
		print("[ERROR] " + text)
	# Success tag
	elif(tag == 's'):
		print("[SUCCESS] " + text)

def save_excel(_FILENAME,_DATA,_HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        sheet.append(_DATA)
        book.save(_FILENAME)
    else: # 새로만드는건
        book = Workbook()
        sheet = book.active
        sheet.title = '시트이름'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        book.save(_FILENAME)

def get_category():
    linkList = []
    titleList = []
    bs4 = get_bs_obejct_by_url('https://delectable.com/categories/')
    divs = bs4.find_all('div', class_='categories-list__section__category')
    for div in divs:
        linkList.append('https://delectable.com' + div.a['href'])
        titleList.append(div.a.get_text().strip())
    return linkList,titleList

if __name__ == "__main__":
    linkList, titleList = get_category()
    f = open("category.txt",'w',encoding='utf8')
    for idx in range(len(linkList)):
        f.write("{}@@@{}\n".format(titleList[idx],linkList[idx]))
    f.close()
    #exit(-1)
    f = open('category.txt')
    lines = f.readlines()

    SCROLL_PAUSE_TIME = 0.1
    driver = webdriver.Chrome('./chromedriver')
    driver.maximize_window()
    for line in lines:
        title, url = line.split('@@@')
        print(" {} 카테고리 파싱중 ".format(title))
        driver.get(url.strip())
        time.sleep(3)
        # Get scroll height
        idx = 1
        while True:
            # Scroll down to bottom
            last_height = driver.execute_script("return document.body.scrollHeight")
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

            while True:
                # Wait to load page
                time.sleep(SCROLL_PAUSE_TIME)

                # Calculate new scroll height and compare with last scroll height
                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    idx += 1
                    break
                else:
                    idx = 1
                last_height = new_height
            if idx == 50:
                break
        time.sleep(3)
        # === save
        linkList = []
        bs4 = BeautifulSoup(driver.page_source, 'lxml')
        divs = bs4.find_all('div', class_='capture capture--feed-capture')
        for div in divs:
            # print(div.find('h1',class_='capture-header__name').get_text())
            if div.a['href'] != '':
                #print('https://delectable.com' + div.a['href'])
                linkList.append('https://delectable.com' + div.a['href'])
        print("총 {}개 수집완료".format(len(linkList)))
        fileTitle = bs4.find('h1',class_='feed__header__titlenopadding').get_text().strip()
        print("파일명은 ",fileTitle,'입니다')
        linkFile = fileTitle
        f = open(linkFile+'.txt','w')
        for link in linkList:
            f.write(link+'\n')
        f.close()

    driver.quit()
