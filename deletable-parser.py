#-*-encoding:utf8-*-
from bs4 import BeautifulSoup
from openpyxl import load_workbook,Workbook
import requests
import os
import shutil

def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        for depth1List in _DATA:
            sheet.append(depth1List)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = '시트이름'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 50
        book.save(_FILENAME)


'''----------------- 환경설정 부분 ---------------------'''
FILENAME = "delectable.xlsx"
HEADER = ['Category','Brand','Name', 'Rating', 'Region', 'Varietal', 'Pairings']

'''---------------------------------------------------'''
# init
#save_excel(FILENAME,None,HEADER)
fileList =os.listdir('./미완/')
for file in fileList:

    CATEGORY = file.split('.')[0]
    urlList = open('./미완/'+file).readlines()
    print(">>> {} 파일 실행중".format(CATEGORY))
    idx = 1
    date2dList = []
    for url in urlList:
        dataList = []
        print("{}번째 : {} 실행중..".format(idx,url.strip()))
        idx +=1
        html = requests.get(url.strip())
        bs4 = BeautifulSoup(html.text,'lxml')


        brand = bs4.find('h2',class_='wine-profile-header__producer').get_text().strip()
        name = bs4.find('h1',class_='wine-profile-header__name').get_text().strip()

        ratingValue = bs4.find("span",itemprop='ratingValue').get_text().strip()
        ratingCount = bs4.find("span",class_='wine-profile-rating__count').get_text().strip()
        proValue =bs4.find_all('div',class_='wine-profile-rating__rating')[1].find("span").get_text().strip()
        proCount = bs4.find_all('div',class_='wine-profile-rating__rating')[1].find("span",class_='wine-profile-rating__count').get_text().strip()
        try:
            regionName = bs4.find('span',class_='wine-profile-region__name').get_text().strip()
        except:
            regionName = '-'
        try:
            varietalName = bs4.find('span',class_='wine-profile-varietal__name').get_text().strip()
        except:
            varietalName = '-'
        try:
            pairings = bs4.find('div',class_='wine-profile-pairings__name').get_text().strip()
        except:
            pairings = '-'
        dataList.append(CATEGORY)
        dataList.append(brand)
        dataList.append(name)
        dataList.append(ratingValue+' '+ratingCount+','+proValue+' '+proCount)
        dataList.append(regionName)
        dataList.append(varietalName)
        dataList.append(pairings)
        date2dList.append(dataList)
    save_excel(FILENAME,date2dList,None)
    shutil.move("./미완/{}".format(file), "./완료/")