#-*-encoding:utf8-*-
from bs4 import BeautifulSoup
from openpyxl import load_workbook,Workbook
import os
from selenium import webdriver

def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        sheet.append(_DATA)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = '시트이름'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 50
        book.save(_FILENAME)

def log(tag, text):
	# Info tag
	if(tag == 'i'):
		print("[INFO] " + text)
	# Error tag
	elif(tag == 'e'):
		print("[ERROR] " + text)
	# Success tag
	elif(tag == 's'):
		print("[SUCCESS] " + text)

if __name__ == "__main__":
    '''----------------- 환경설정 부분 ---------------------'''
    FILENAME = input(">>> 저장 엑셀파일이름을 확장자포함(.xlsx) 적어주세요 :: ").strip()
    HEADER = ['Winery', 'Name', 'Region', 'Country', 'Rating', 'Summary']
    # init
    save_excel(FILENAME, None, HEADER)
    driver = webdriver.Chrome('./chromedriver')
    driver.maximize_window()
    driver.get('https://www.vivino.com')
    f2 = open('vivino_except_list.txt', 'a')
    while True:
        inputStr = input(">>> 추출 원하는 페이지에 들어가서 아무키나 눌러주세요 (종료는 0) :").strip()
        if inputStr == '0':
            break
        dataList = []
        driver.switch_to.window(driver.window_handles[1])
        bs4 = BeautifulSoup(driver.page_source, 'lxml')
        #
        try:
            name = bs4.find('span', class_='wine-page__header__information__details__name__vintage').get_text().strip()
        except:  # 새로운페이지
            name = '-'
            print("\t>>> Move Except File")
            f2.write(driver.current_url+'\n')
            continue
        try:
            winery = bs4.find('a', class_='wine-page__header__information__details__name__winery').get_text().strip()
        except:
            winery = '-'
        try:
            region = bs4.find('a',
                              class_='wine-page__header__information__details__location__region').get_text().strip()
        except:
            region = '-'
        try:
            country = bs4.find("a",
                               class_='wine-page__header__information__details__location__country').get_text().strip()
        except:
            country = '-'
        try:
            averLabel = bs4.find('div',
                                 class_='wine-page__header__information__details__average-rating__label').get_text().strip()
        except:
            averLabel = '-'
        try:
            averValue = bs4.find('div',
                                 class_='wine-page__header__information__details__average-rating__value__number').get_text().strip()
        except:
            averValue = '-'
        divs = bs4.find_all('div', class_='wine-page__summary__item')
        summary = ""
        for div in divs:
            summaryHeader = div.find('div', class_='wine-page__summary__item__header').get_text().strip()
            summary += summaryHeader + ":"
            isVoid = True
            for a in div.find_all('a'):
                summary += a.get_text().strip() + ','
                isVoid = False
            for a in div.find_all('span'):
                summary += a.get_text().strip() + ','
                isVoid = False
            if isVoid:
                summary += div.find('div', class_='wine-page__summary__item__content').get_text().strip() + ','
            summary = summary[:-1] + '\n'

        dataList.append(winery)
        dataList.append(name)
        dataList.append(region)
        dataList.append(country)
        dataList.append(averLabel + ' ' + averValue)
        dataList.append(summary)
        save_excel(FILENAME, dataList, None)
        log('s', name + " 저장 완료")
    driver.quit()
    f2.close()